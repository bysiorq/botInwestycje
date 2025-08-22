# ────────────────────────── etapy_bot.py (2025-08 • FIX: brak duplikatów panelu + stabilne % i zapisy) ──────────────────────────
# Single-message UI jak BotFather. Stabilne na webhooku/wielu replikach:
# - Stan per user w /data/state/<uid>.json (project, stage_code, await, date)
# - Callbacki niosą stage_code (S1..S7) → brak zależności od ulotnego user_data
# - Każda zmiana zapisuje do Excela i renderuje widok; jeśli treść się nie zmieni, NIE tworzymy nowej wiadomości
# - Nazwy arkuszy Excela sanityzowane (max 31, bez : \ / ? * [ ])

import os
import re
import json
import logging
import calendar as cal
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    from office365.sharepoint.client_context import ClientContext
    from office365.runtime.auth.client_credential import ClientCredential
except ModuleNotFoundError:
    ClientContext = ClientCredential = None

from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup, BotCommand,
)
from telegram.ext import (
    ApplicationBuilder, Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, ContextTypes, filters, ConversationHandler,
)
from telegram.error import BadRequest

import tempfile
import portalocker

# ──────────────────── konfiguracja ────────────────────
load_dotenv()
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
WEBHOOK_URL = os.getenv("WEBHOOK_URL", "").rstrip("/")
PORT = int(os.getenv("PORT", 8080))
DATA_DIR = os.getenv("DATA_DIR", ".")
os.makedirs(DATA_DIR, exist_ok=True)

EXCEL_FILE = os.path.join(DATA_DIR, "projects.xlsx")
LOCK_FILE = os.path.join(DATA_DIR, "projects.lock")
STATE_DIR = os.path.join(DATA_DIR, "state")
os.makedirs(STATE_DIR, exist_ok=True)

PROJECTS_SHEET = "__Projects"
PROJECTS_HEADERS = ["Project", "Active", "Finished", "CreatedAt"]

STAGES = [
    {"code": "S1", "name": "Etap 1"},
    {"code": "S2", "name": "Etap 2"},
    {"code": "S3", "name": "Etap 3"},
    {"code": "S4", "name": "Etap 4"},
    {"code": "S5", "name": "Etap 5"},
    {"code": "S6", "name": "Etap 6"},
    {"code": "S7", "name": "Prace dodatkowe"},
]
CODE2NAME = {x["code"]: x["name"] for x in STAGES}
NAME2CODE = {x["name"]: x["code"] for x in STAGES}

STAGE_HEADERS = [
    "Stage", "Percent", "ToFinish", "Notes", "Finished",
    "LastUpdated", "Photos", "LastEditor", "LastEditorId",
]

DATE_PICK = 10

# ──────────────────── helpers ────────────────────
def today_str() -> str: return datetime.now().strftime("%d.%m.%Y")
def to_ddmmyyyy(d: date) -> str: return d.strftime("%d.%m.%Y")

def _atomic_save_wb(wb: Workbook, path: str) -> None:
    fd, tmp_path = tempfile.mkstemp(dir=os.path.dirname(path), suffix=".tmp"); os.close(fd)
    wb.save(tmp_path); os.replace(tmp_path, path)

def _atomic_save_json(path: str, obj: dict) -> None:
    fd, tmp_path = tempfile.mkstemp(dir=os.path.dirname(path), suffix=".tmp"); os.close(fd)
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False)
    os.replace(tmp_path, path)

def _with_lock(fn, *args, **kwargs):
    with portalocker.Lock(LOCK_FILE, timeout=30):
        return fn(*args, **kwargs)

# ──────────────────── stan użytkownika (trwały) ────────────────────
def _state_path(uid: int) -> str:
    return os.path.join(STATE_DIR, f"{uid}.json")

def load_user_state(uid: int) -> dict:
    path = _state_path(uid)
    if not os.path.exists(path): return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_user_state(uid: int, data: dict) -> None:
    try:
        _atomic_save_json(_state_path(uid), data)
    except Exception:
        pass

def sync_in(update_or_ctx, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = (update_or_ctx.effective_user.id if isinstance(update_or_ctx, Update)
           else update_or_ctx.callback_query.from_user.id)
    state = load_user_state(uid)
    if state:
        for k in ["date", "project", "stage_code", "await"]:
            if k in state:
                context.user_data[k] = state[k]
    return uid

def sync_out(uid: int, context: ContextTypes.DEFAULT_TYPE) -> None:
    data = {}
    for k in ["date", "project", "stage_code", "await", "sticky_id"]:
        if k in context.user_data:
            data[k] = context.user_data[k]
    save_user_state(uid, data)

# ──────────────────── Excel ────────────────────
def _sheet_title(project_name: str) -> str:
    bad = set(':\/?*[]')
    title = "".join(('·' if ch in bad else ch) for ch in project_name)
    title = title[:31] if len(title) > 31 else title
    return title or "Projekt"

def open_wb() -> Workbook:
    if os.path.exists(EXCEL_FILE):
        return load_workbook(EXCEL_FILE)
    wb = Workbook()
    ws = wb.active; ws.title = PROJECTS_SHEET; ws.append(PROJECTS_HEADERS)
    _atomic_save_wb(wb, EXCEL_FILE)
    return wb

def ensure_projects_sheet(wb: Workbook) -> Worksheet:
    if PROJECTS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(PROJECTS_SHEET, index=0); ws.append(PROJECTS_HEADERS)
    else:
        ws = wb[PROJECTS_SHEET]
        if ws.max_row < 1: ws.append(PROJECTS_HEADERS)
    return ws

def list_projects(active_only: bool = True) -> List[Dict[str, str]]:
    def _read():
        wb = open_wb(); ws = ensure_projects_sheet(wb)
        out = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            prj = {
                "name": row[0],
                "active": (str(row[1]).lower() != "false"),
                "finished": (str(row[2]).lower() == "true"),
                "created": row[3],
            }
            if not active_only or prj["active"]: out.append(prj)
        return out
    return _with_lock(_read)

def add_project(name: str) -> None:
    name = name.strip()
    if not name: return
    def _upd():
        wb = open_wb(); ws = ensure_projects_sheet(wb)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == name: return
        ws.append([name, True, False, datetime.now().isoformat()])
        title = _sheet_title(name)
        if title not in wb.sheetnames:
            ws2 = wb.create_sheet(title); ws2.append(STAGE_HEADERS)
            for st in STAGES: ws2.append([st["name"], "", "", "", "-", "", "", "", ""])
        _atomic_save_wb(wb, EXCEL_FILE)
    _with_lock(_upd)

def set_project_active(name: str, active: bool) -> None:
    def _upd():
        wb = open_wb(); ws = ensure_projects_sheet(wb)
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value == name:
                ws.cell(r, 2, True if active else False); break
        _atomic_save_wb(wb, EXCEL_FILE)
    _with_lock(_upd)

def set_project_finished(name: str, finished: bool) -> None:
    def _upd():
        wb = open_wb(); ws = ensure_projects_sheet(wb)
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value == name:
                ws.cell(r, 3, True if finished else False); break
        _atomic_save_wb(wb, EXCEL_FILE)
    _with_lock(_upd)

def ensure_project_sheet(name: str) -> Worksheet:
    wb = open_wb()
    title = _sheet_title(name)
    if title not in wb.sheetnames:
        ws2 = wb.create_sheet(title); ws2.append(STAGE_HEADERS)
        for st in STAGES: ws2.append([st["name"], "", "", "", "-", "", "", "", ""])
        _atomic_save_wb(wb, EXCEL_FILE)
    return wb[title]

def read_stage(project: str, stage_name: str) -> Dict[str, str]:
    def _read():
        ws = ensure_project_sheet(project)
        headers = [c.value for c in ws[1]]; idx = {h: i for i, h in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[idx["Stage"]] == stage_name:
                return {
                    "Stage": row[idx["Stage"]],
                    "Percent": row[idx["Percent"]] if row[idx["Percent"]] is not None else "",
                    "ToFinish": row[idx["ToFinish"]] or "",
                    "Notes": row[idx["Notes"]] or "",
                    "Finished": row[idx["Finished"]] or "-",
                    "LastUpdated": row[idx["LastUpdated"]] or "",
                    "Photos": row[idx["Photos"]] or "",
                    "LastEditor": row[idx["LastEditor"]] or "",
                    "LastEditorId": row[idx["LastEditorId"]] or "",
                }
        ws2 = ensure_project_sheet(project); ws2.append([stage_name, "", "", "", "-", "", "", "", ""])
        _atomic_save_wb(ws2.parent, EXCEL_FILE)
        return {"Stage": stage_name, "Percent": "", "ToFinish": "", "Notes": "", "Finished": "-", "LastUpdated": "", "Photos": "", "LastEditor": "", "LastEditorId": ""}
    return _with_lock(_read)

def update_stage(project: str, stage_name: str, updates: Dict[str, str], editor_name: str, editor_id: int) -> None:
    allowed = set(STAGE_HEADERS) - {"Stage"}
    for k in updates.keys():
        if k not in allowed: raise ValueError(f"Unsupported field: {k}")
    def _upd():
        wb = open_wb(); ws = ensure_project_sheet(project)
        headers = [c.value for c in ws[1]]; hidx = {h: i+1 for i, h in enumerate(headers)}
        target_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, hidx["Stage"]).value == stage_name:
                target_row = r; break
        if target_row is None:
            target_row = ws.max_row + 1; ws.append([stage_name] + [""]*(len(headers)-1))
        for k, v in updates.items():
            ws.cell(target_row, hidx[k], v)
        # Dodaj sekundy, by łatwiej różnicować treść panelu
        ws.cell(target_row, hidx["LastUpdated"], datetime.now().strftime("%d.%m.%Y %H:%M:%S"))
        ws.cell(target_row, hidx["LastEditor"], editor_name)
        ws.cell(target_row, hidx["LastEditorId"], str(editor_id))
        _atomic_save_wb(wb, EXCEL_FILE)
    _with_lock(_upd)

# ──────────────────── UI helpers ────────────────────
async def safe_answer(q, text: Optional[str] = None, show_alert: bool = False):
    try:
        if text is not None: await q.answer(text=text, show_alert=show_alert)
        else: await q.answer()
    except BadRequest:
        pass
    except Exception:
        pass

async def sticky_set(update_or_ctx, context: ContextTypes.DEFAULT_TYPE, text: str, reply_markup: Optional[InlineKeyboardMarkup] = None):
    """Edytuje istniejący panel; jeśli identyczny – NIC nie robi; nową wiadomość wysyła tylko gdy nie ma sticky_id albo
    gdy edycja jest niemożliwa (np. 'message to edit not found')."""
    chat = update_or_ctx.effective_chat if isinstance(update_or_ctx, Update) else update_or_ctx.callback_query.message.chat
    chat_id = chat.id
    sticky_id = context.user_data.get("sticky_id")
    if sticky_id:
        try:
            await context.bot.edit_message_text(
                chat_id=chat_id, message_id=sticky_id, text=text,
                reply_markup=reply_markup, parse_mode="Markdown", disable_web_page_preview=True
            )
            return
        except BadRequest as e:
            msg = str(e).lower()
            # NIE twórz nowej wiadomości, jeśli treść się nie zmieniła
            if "message is not modified" in msg:
                return
            # Jeśli edycja niemożliwa (skasowane itp.) – przejdź do wysłania nowej
            if not any(s in msg for s in [
                "message to edit not found",
                "message identifier is not specified",
                "chat not found",
                "message can't be edited",
            ]):
                # Inne błędy – też nie duplikuj panelu
                return
        except Exception:
            # Nie duplikuj przy nieznanym błędzie
            return
    # Brak sticky_id albo poprzednia wiadomość nie istnieje → wyślij nową i zapamiętaj id
    m = await context.bot.send_message(chat_id, text, reply_markup=reply_markup, parse_mode="Markdown", disable_web_page_preview=True)
    context.user_data["sticky_id"] = m.message_id
    # zapisz sticky_id trwale (przydatne przy replikach)
    uid = update_or_ctx.effective_user.id if isinstance(update_or_ctx, Update) else update_or_ctx.callback_query.from_user.id
    save_user_state(uid, {**load_user_state(uid), "sticky_id": m.message_id, "date": context.user_data.get("date"),
                          "project": context.user_data.get("project"), "stage_code": context.user_data.get("stage_code"),
                          "await": context.user_data.get("await")})

def banner_await(context: ContextTypes.DEFAULT_TYPE) -> str:
    aw = context.user_data.get("await") or {}
    if not aw: return ""
    names = {"project_name": "Nazwa inwestycji", "todo": "Do dokończenia", "notes": "Notatki", "percent": "% ukończenia", "photo": "Zdjęcie"}
    proj = context.user_data.get("project") or ""
    scode = context.user_data.get("stage_code") or ""
    sname = CODE2NAME.get(scode, "")
    where = f" (inwestycja: {proj}" + (f" | {sname}" if sname else "") + ")"
    return f"✍️ *Oczekuję na:* {names.get(aw.get('field'), aw.get('field'))}{where}. Wyślij teraz.\n"

def projects_menu_text(context: ContextTypes.DEFAULT_TYPE) -> str:
    ds = context.user_data.get("date", today_str())
    out = []
    b = banner_await(context); 
    if b: out.append(b)
    out.append(f"🏗️ *Inwestycje*  |  📅 {ds}\n")
    if not list_projects(active_only=True):
        out.append("Brak inwestycji. Dodaj pierwszą 👇")
    return "\n".join(out)

def projects_menu_kb(context: ContextTypes.DEFAULT_TYPE) -> InlineKeyboardMarkup:
    ds = context.user_data.get("date", today_str())
    projs = list_projects(active_only=True)
    aw = context.user_data.get("await") or {}
    adding = (aw.get("mode") == "text" and aw.get("field") == "project_name")
    def mark(lbl, on): return f"{'●' if on else '○'} {lbl}"
    rows = [[InlineKeyboardButton(f"📅 Data: {ds}", callback_data="date:open")]]
    for i, p in enumerate(projs):
        rows.append([InlineKeyboardButton(f"🏗️ {p['name']}", callback_data=f"proj:open:{i}")])
    rows.append([InlineKeyboardButton(mark("➕ Dodaj inwestycję", adding), callback_data="proj:add")])
    rows.append([InlineKeyboardButton("🗄 Archiwum", callback_data="proj:arch")])
    return InlineKeyboardMarkup(rows)

def _percent_preview_for_project(project: str) -> str:
    try:
        ws = ensure_project_sheet(project)
        headers = [c.value for c in ws[1]]; hidx = {h: i+1 for i, h in enumerate(headers)}
        vals = {}
        for r in range(2, ws.max_row + 1):
            st = ws.cell(r, hidx["Stage"]).value
            p = ws.cell(r, hidx["Percent"]).value
            vals[st] = "-" if (p in ("", None)) else f"{int(p)}%" if str(p).isdigit() else str(p)
        parts = []
        for st in STAGES:
            parts.append(f"{st['name'].split()[-1]} {vals.get(st['name'], '-')}")
        return " | ".join(parts)
    except Exception:
        return "-"

def project_panel_text(context: ContextTypes.DEFAULT_TYPE) -> str:
    proj = context.user_data.get("project")
    out = []
    b = banner_await(context)
    if b: out.append(b)
    out.append(f"🏗️ *{proj}*")
    out.append(f"📊 Postęp etapów: {_percent_preview_for_project(proj)}\n")
    out.append("👇 Wybierz etap. Otwarte zadania:")
    for st in STAGES:
        data = read_stage(proj, st["name"])
        tf = (data["ToFinish"] or "").strip()
        p = data["Percent"]
        ptxt = f" (📊 {int(p)}%)" if str(p).isdigit() else ""
        if tf:
            prev = tf if len(tf) <= 60 else tf[:57] + "…"
            out.append(f"• {st['name']}{ptxt}: 🔧 {prev}")
    return "\n".join(out)

def project_panel_kb(context: ContextTypes.DEFAULT_TYPE) -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton("Etap 1", callback_data="stage:open:S1"),
         InlineKeyboardButton("Etap 2", callback_data="stage:open:S2")],
        [InlineKeyboardButton("Etap 3", callback_data="stage:open:S3"),
         InlineKeyboardButton("Etap 4", callback_data="stage:open:S4")],
        [InlineKeyboardButton("Etap 5", callback_data="stage:open:S5"),
         InlineKeyboardButton("Etap 6", callback_data="stage:open:S6")],
        [InlineKeyboardButton("Prace dodatkowe", callback_data="stage:open:S7")],
        [InlineKeyboardButton("✅ Oznacz zakończoną", callback_data="proj:finish"),
         InlineKeyboardButton("📦 Archiwizuj/Przywróć", callback_data="proj:toggle_active")],
        [InlineKeyboardButton("↩️ Wstecz", callback_data="nav:home")],
    ]
    return InlineKeyboardMarkup(rows)

def stage_panel_text(context: ContextTypes.DEFAULT_TYPE) -> str:
    proj = context.user_data.get("project")
    scode = context.user_data.get("stage_code")
    sname = CODE2NAME.get(scode, "")
    data = read_stage(proj, sname)
    out = []
    b = banner_await(context)
    if b: out.append(b)
    out.extend([
        f"🏗️ *{proj}*  →  {sname}",
        "",
        f"📊 % ukończenia: {data['Percent'] if data['Percent'] != '' else '-'}",
        f"🔧 Do dokończenia:\n{data['ToFinish'] or '-'}",
        f"📝 Notatki:\n{data['Notes'] or '-'}",
        f"🖼 Zdjęcia: {len((data['Photos'] or '').split()) if (data['Photos'] or '').strip() else 0}",
        f"⏱ Ostatnia zmiana: {data['LastUpdated'] or '-'}  |  👤 {data['LastEditor'] or '-'}",
        "",
        "Wybierz działanie poniżej 👇",
    ])
    return "\n".join(out)

def stage_panel_kb(context: ContextTypes.DEFAULT_TYPE) -> InlineKeyboardMarkup:
    aw = context.user_data.get("await") or {}
    active_key = None
    if aw:
        if aw.get("mode") == "text" and aw.get("field") in {"todo", "notes", "percent"}: active_key = aw.get("field")
        if aw.get("mode") == "photo": active_key = "photo"
    scode = context.user_data.get("stage_code", "")
    def mark(lbl, key): return f"{'●' if active_key == key else '○'} {lbl}"
    rows = [
        [InlineKeyboardButton(mark("🔧 Do dokończenia", "todo"), callback_data="stage:set:todo"),
         InlineKeyboardButton(mark("📝 Notatki", "notes"), callback_data="stage:set:notes")],
        [InlineKeyboardButton(mark("📊 % (0/25/50/75/90/100)", "percent"), callback_data=f"stage:set:percent:{scode}"),
            InlineKeyboardButton(mark("📸 Dodaj zdjęcie", "photo"), callback_data="stage:add_photo")],
        [InlineKeyboardButton("🧹 Wyczyść Do dokończenia", callback_data=f"stage:clear:todo:{scode}"),
         InlineKeyboardButton("🧹 Wyczyść Notatki", callback_data=f"stage:clear:notes:{scode}")],
        [InlineKeyboardButton("💾 Zapisz zmiany", callback_data=f"stage:save:{scode}")],
        [InlineKeyboardButton("↩️ Wstecz", callback_data="proj:back")],
    ]
    return InlineKeyboardMarkup(rows)

def percent_kb(stage_code: str) -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton("0%", callback_data=f"pct:{stage_code}:0"),
         InlineKeyboardButton("25%", callback_data=f"pct:{stage_code}:25"),
         InlineKeyboardButton("50%", callback_data=f"pct:{stage_code}:50")],
        [InlineKeyboardButton("75%", callback_data=f"pct:{stage_code}:75"),
         InlineKeyboardButton("90%", callback_data=f"pct:{stage_code}:90"),
         InlineKeyboardButton("100%", callback_data=f"pct:{stage_code}:100")],
        [InlineKeyboardButton("✍️ Wpisz ręcznie", callback_data=f"pct:{stage_code}:manual")],
        [InlineKeyboardButton("↩️ Wróć", callback_data="pct:back")],
    ]
    return InlineKeyboardMarkup(rows)

def month_kb(year: int, month: int) -> InlineKeyboardMarkup:
    month_name = cal.month_name[month]; days = cal.monthcalendar(year, month)
    rows = [[InlineKeyboardButton(f"{month_name} {year}", callback_data="noop")]]
    rows.append([InlineKeyboardButton(x, callback_data="noop") for x in ["Pn","Wt","Śr","Cz","Pt","So","Nd"]])
    for week in days:
        r = []
        for d in week:
            if d == 0: r.append(InlineKeyboardButton(" ", callback_data="noop"))
            else:
                ds = to_ddmmyyyy(date(year, month, d))
                r.append(InlineKeyboardButton(str(d), callback_data=f"day:{ds}"))
        rows.append(r)
    prev_month = (date(year, month, 1) - timedelta(days=1))
    next_month = (date(year, month, cal.monthrange(year, month)[1]) + timedelta(days=1))
    rows.append([
        InlineKeyboardButton("« Poprzedni", callback_data=f"cal:{prev_month.year}-{prev_month.month:02d}"),
        InlineKeyboardButton("Dziś", callback_data=f"day:{today_str()}"),
        InlineKeyboardButton("Następny »", callback_data=f"cal:{next_month.year}-{next_month.month:02d}"),
    ])
    rows.append([InlineKeyboardButton("↩️ Wstecz", callback_data="nav:home")])
    return InlineKeyboardMarkup(rows)

# ──────────────────── renderery ────────────────────
async def render_home(update_or_ctx, context: ContextTypes.DEFAULT_TYPE):
    sync_in(update_or_ctx, context)
    await sticky_set(update_or_ctx, context, projects_menu_text(context), projects_menu_kb(context))

async def render_project(update_or_ctx, context: ContextTypes.DEFAULT_TYPE):
    sync_in(update_or_ctx, context)
    await sticky_set(update_or_ctx, context, project_panel_text(context), project_panel_kb(context))

async def render_stage(update_or_ctx, context: ContextTypes.DEFAULT_TYPE):
    sync_in(update_or_ctx, context)
    await sticky_set(update_or_ctx, context, stage_panel_text(context), stage_panel_kb(context))

# ──────────────────── Handlers ────────────────────
async def start_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = sync_in(update, context)
    context.user_data.clear()
    context.user_data["date"] = today_str()
    sync_out(uid, context)
    await render_home(update, context)

async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = sync_in(update, context); sync_out(uid, context)
    text = (
        "🤖 *Pomoc – Inwestycje*\n"
        "• /start – lista inwestycji, dodawanie, archiwum.\n"
        "• W projekcie → Etap → edycja pól. Zmiany zapisują się do Excela i natychmiast widać w panelu.\n"
        "• Kropki ○/● pokazują, że czekam na tekst/zdjęcie.\n"
        "• Stan sesji jest trwały (działa stabilnie na webhooku/skalowaniu).\n"
    )
    await sticky_set(update, context, text, InlineKeyboardMarkup([[InlineKeyboardButton("↩️ Wstecz", callback_data="nav:home")]]))

# --- data ---
async def date_open_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = sync_in(update, context); await safe_answer(update.callback_query)
    now = datetime.now()
    await sticky_set(update, context, "📅 Wybierz datę (informacyjnie):", month_kb(now.year, now.month))
    sync_out(uid, context)
    return DATE_PICK

async def calendar_nav_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = sync_in(update, context); q = update.callback_query; await safe_answer(q); data = q.data
    if data.startswith("cal:"):
        y, m = map(int, data.split(":")[1].split("-"))
        await sticky_set(update, context, "📅 Wybierz datę:", month_kb(y, m)); sync_out(uid, context); return DATE_PICK
    if data.startswith("day:"):
        ds = data.split(":")[1]; context.user_data["date"] = ds; sync_out(uid, context)
        await render_home(update, context); return ConversationHandler.END
    sync_out(uid, context); return DATE_PICK

# --- projekty / archiwum ---
def _render_archive_kb(context: ContextTypes.DEFAULT_TYPE) -> InlineKeyboardMarkup:
    projs = list_projects(active_only=False)
    context.user_data["arch_names"] = [p["name"] for p in projs]
    rows = []
    for i, p in enumerate(projs):
        state = "🟢" if p["active"] else "⚪️"
        rows.append([InlineKeyboardButton(f"{state} {p['name']}", callback_data=f"arch:tog:{i}")])
    rows.append([InlineKeyboardButton("↩️ Wstecz", callback_data="nav:home")])
    return InlineKeyboardMarkup(rows)

async def projects_router(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = sync_in(update, context); q = update.callback_query; await safe_answer(q); data = q.data

    if data == "nav:home":
        sync_out(uid, context); await render_home(update, context); return

    if data == "proj:add":
        context.user_data["await"] = {"mode": "text", "field": "project_name"}; sync_out(uid, context)
        await render_home(update, context); return

    if data == "proj:arch":
        await sticky_set(update, context, "🗄 *Archiwum / Aktywne* (kliknij aby przełączyć):", _render_archive_kb(context)); sync_out(uid, context); return

    if data.startswith("arch:tog:"):
        idx = int(data.split(":")[2]); names = context.user_data.get("arch_names", [])
        if 0 <= idx < len(names):
            allp = {p["name"]: p for p in list_projects(active_only=False)}
            cur = allp.get(names[idx])
            if cur: set_project_active(names[idx], not cur["active"])
        await sticky_set(update, context, "🗄 *Archiwum / Aktywne* (kliknij aby przełączyć):", _render_archive_kb(context)); sync_out(uid, context); return

    if data.startswith("proj:open:"):
        idx = int(data.split(":")[2]); projs = list_projects(active_only=True)
        if 0 <= idx < len(projs):
            context.user_data["project"] = projs[idx]["name"]; context.user_data.pop("await", None); sync_out(uid, context)
            await render_project(update, context)
        else:
            await render_home(update, context)
        return

    if data == "proj:finish":
        proj = context.user_data.get("project")
        if not proj: sync_out(uid, context); await render_home(update, context); return
        set_project_finished(proj, True)
        await sticky_set(update, context, f"🎉 *{proj}* oznaczono jako zakończoną. 💪", InlineKeyboardMarkup([[InlineKeyboardButton("↩️ Wróć", callback_data="nav:home")]]))
        sync_out(uid, context); return

    if data == "proj:toggle_active":
        proj = context.user_data.get("project")
        if proj:
            allp = {p["name"]: p for p in list_projects(active_only=False)}
            cur = allp.get(proj)
            if cur: set_project_active(proj, not cur["active"])
        sync_out(uid, context); await render_home(update, context); return

# --- panel etapu ---
async def stage_router(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = sync_in(update, context); q = update.callback_query; await safe_answer(q); data = q.data
    proj = context.user_data.get("project")

    if data.startswith("stage:open:"):
        scode = data.split(":")[2]
        if scode not in CODE2NAME: sync_out(uid, context); await render_project(update, context); return
        context.user_data["stage_code"] = scode; context.user_data.pop("await", None); sync_out(uid, context)
        await render_stage(update, context); return

    if data == "stage:set:todo":
        context.user_data["await"] = {"mode": "text", "field": "todo"}; sync_out(uid, context)
        await render_stage(update, context); return
    if data == "stage:set:notes":
        context.user_data["await"] = {"mode": "text", "field": "notes"}; sync_out(uid, context)
        await render_stage(update, context); return

    if data.startswith("stage:set:percent:"):
        scode = data.split(":")[3]
        await sticky_set(update, context, "📊 Ustaw % ukończenia:", percent_kb(scode)); sync_out(uid, context); return

    if data.startswith("stage:clear:"):
        _, _, field, scode = data.split(":")
        sname = CODE2NAME.get(scode, "")
        if proj and sname:
            update_stage(proj, sname, {"ToFinish" if field == "todo" else "Notes": ""}, q.from_user.first_name, q.from_user.id)
        await safe_answer(q, "Wyczyszczono ✅"); sync_out(uid, context); await render_stage(update, context); return

    if data.startswith("stage:save:"):
        scode = data.split(":")[2]; sname = CODE2NAME.get(scode, "")
        if proj and sname:
            update_stage(proj, sname, {}, q.from_user.first_name, q.from_user.id)
        await safe_answer(q, "Zapisano ✅"); sync_out(uid, context); await render_stage(update, context); return

    if data == "proj:back":
        context.user_data.pop("await", None); sync_out(uid, context); await render_project(update, context); return

    if data == "stage:add_photo":
        context.user_data["await"] = {"mode": "photo", "field": "photo"}; sync_out(uid, context)
        await render_stage(update, context); return

# --- procenty ---
async def percent_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = sync_in(update, context); q = update.callback_query; await safe_answer(q)
    data = q.data
    proj = context.user_data.get("project")

    if data == "pct:back":
        sync_out(uid, context); await render_stage(update, context); return

    if data.startswith("pct:"):
        try:
            _, scode, val = data.split(":")
        except ValueError:
            return
        sname = CODE2NAME.get(scode, "")
        if val == "manual":
            context.user_data["await"] = {"mode": "text", "field": "percent"}; context.user_data["stage_code"] = scode; sync_out(uid, context)
            await render_stage(update, context); return
        pct = None
        try: pct = int(val)
        except Exception: pass
        if proj and sname and pct is not None:
            update_stage(proj, sname, {"Percent": pct}, q.from_user.first_name, q.from_user.id)
            await safe_answer(q, "Ustawiono % ✅")
        sync_out(uid, context); await render_stage(update, context); return

# --- tekstowe wejścia ---
async def text_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = sync_in(update, context)
    txt = (update.message.text or "").strip()
    try: await update.message.delete()
    except Exception: pass

    aw = context.user_data.get("await") or {}
    mode = aw.get("mode"); field = aw.get("field")

    if mode == "text" and field == "project_name":
        if txt: add_project(txt)
        context.user_data.pop("await", None); sync_out(uid, context)
        await render_home(update, context); return

    if mode != "text":
        sync_out(uid, context); return

    proj = context.user_data.get("project")
    scode = context.user_data.get("stage_code")
    sname = CODE2NAME.get(scode or "", "")
    if not proj or not sname:
        context.user_data.pop("await", None); sync_out(uid, context)
        await render_home(update, context); return

    if field == "todo":
        update_stage(proj, sname, {"ToFinish": txt}, update.effective_user.first_name, update.effective_user.id)
    elif field == "notes":
        update_stage(proj, sname, {"Notes": txt}, update.effective_user.first_name, update.effective_user.id)
    elif field == "percent":
        if not re.fullmatch(r"\d{1,3}", txt):
            await sticky_set(update, context, "📊 Wpisz liczbę 0-100:", percent_kb(scode)); return
        val = int(txt)
        if not (0 <= val <= 100):
            await sticky_set(update, context, "📊 Zakres 0-100:", percent_kb(scode)); return
        update_stage(proj, sname, {"Percent": val}, update.effective_user.first_name, update.effective_user.id)

    context.user_data.pop("await", None); sync_out(uid, context)
    await render_stage(update, context)

# --- zdjęcia ---
async def photo_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = sync_in(update, context)
    aw = context.user_data.get("await") or {}
    if aw.get("mode") != "photo":
        sync_out(uid, context); return
    proj = context.user_data.get("project")
    scode = context.user_data.get("stage_code"); sname = CODE2NAME.get(scode or "", "")
    if not proj or not sname:
        context.user_data.pop("await", None); sync_out(uid, context); return
    try:
        file_id = update.message.photo[-1].file_id
    except Exception:
        try: await update.message.delete()
        except Exception: pass
        return
    data = read_stage(proj, sname)
    photos = (data["Photos"] or "").split(); photos.append(file_id); photos = photos[-200:]
    update_stage(proj, sname, {"Photos": " ".join(photos)}, update.effective_user.first_name, update.effective_user.id)
    try: await update.message.delete()
    except Exception: pass
    context.user_data.pop("await", None); sync_out(uid, context)
    await render_stage(update, context)

# --- cancel / errors ---
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if context.user_data.get("sticky_id"):
            await context.bot.delete_message(update.effective_chat.id, context.user_data.get("sticky_id"))
    except Exception:
        pass
    await update.effective_chat.send_message("Anulowano.")
    context.user_data.clear()
    return ConversationHandler.END

async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    err = context.error
    if isinstance(err, BadRequest) and ("query is too old" in str(err).lower() or "query is not found" in str(err).lower()):
        return
    logging.exception("Unhandled exception: %s", err)

# ──────────────────── PTB Application ────────────────────
async def on_startup(app: Application) -> None:
    await app.bot.set_my_commands([
        BotCommand("start", "Otwórz panel inwestycji"),
        BotCommand("help", "Pomoc"),
    ])

def build_app() -> Application:
    app = ApplicationBuilder().token(TELEGRAM_TOKEN).post_init(on_startup).build()
    app.add_handler(CommandHandler("start", start_cmd))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("cancel", cancel))

    app.add_handler(CallbackQueryHandler(date_open_cb, pattern=r"^date:open$"))
    app.add_handler(CallbackQueryHandler(calendar_nav_cb, pattern=r"^(cal:\d{4}-\d{2}|day:\d{2}\.\d{2}\.\d{4})$"))

    app.add_handler(CallbackQueryHandler(projects_router, pattern=r"^(nav:home|proj:add|proj:arch|arch:tog:\d+|proj:open:\d+|proj:finish|proj:toggle_active)$"))

    app.add_handler(CallbackQueryHandler(stage_router, pattern=r"^(stage:open:S[1-7]|stage:set:(todo|notes)|stage:set:percent:S[1-7]|stage:clear:(todo|notes):S[1-7]|stage:save:S[1-7]|proj:back|stage:add_photo)$"))
    app.add_handler(CallbackQueryHandler(percent_cb, pattern=r"^(pct:(S[1-7]):(\d+|manual)|pct:back)$"))

    app.add_handler(MessageHandler(filters.PHOTO, photo_input))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_input))

    app.add_error_handler(error_handler)
    return app

# ──────────────────── main ────────────────────
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s")
    if not TELEGRAM_TOKEN:
        raise SystemExit("Brak TELEGRAM_TOKEN w env.")
    app = build_app()
    if WEBHOOK_URL:
        app.run_webhook(listen="0.0.0.0", port=PORT, url_path=TELEGRAM_TOKEN, webhook_url=f"{WEBHOOK_URL}/{TELEGRAM_TOKEN}")
    else:
        app.run_polling(allowed_updates=Update.ALL_TYPES)
